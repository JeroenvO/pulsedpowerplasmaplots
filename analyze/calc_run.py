"""
Convert all output of a run to a pickle and xlsx file. These files can be used for plotting data.

In one run, multiple measurements are made. Each measurement can contain multiple scope spectra.

Parse log excel file
Get ozone concentration
Get scope plots

Ocean optics HR2000 -> SpectraSuite -> CSV -> spectrum_parse.{a,b,c} ↴
LeCroy WaveAce 224 -> EasyScope 	-> CSV -> scope_parse.{a,b,c,d}  ↴
Manual written log 					-> XLSX -> calc_run.py			-|-> data.pkl & data.xlsx

Make list of dicts. Each dict is one measurement and contains all relevant input and output values.
Jeroen van Oorschot 2017-2018, Eindhoven University of Technology
jeroen@jjvanoorschot.nl
"""
import os
import pickle
import glob

import numpy as np
from openpyxl.reader.excel import load_workbook, Workbook

from analyze.scope_parse.c_get_lines import get_vol_cur_single, get_vol_cur_multiple
from analyze.scope_parse.d_calc import calc_output
from analyze.scope_parse.e_average import calc_output_avg
from analyze.spectrum_parse.c_concentration import ozone_concentration, ozone_ppm
from visualize.helpers.data import sort_data
from analyze.defines import *

# Folder with measuremtn. Folder has to contain:
# * log.txt
# * spect/m00000.txt  with spectral data
# * scope/000.csv with 000=used input voltage
# See bottom of file for run params.


def calc_run(run_dir,
             reactor,
             spect_dir='spect',
             scope_dir='scope',
             scope_multiple = False,
             log_file='log.xlsx',
             scope_file_name_index=0,
             meas=SHORT_MEAS_LEN,
             current_scaling=0.5,
             delay=0,
             voltage_offset=None,
             current_offset=None,
             waveform_loose_stability=False):
    """
    Calculate all parameters for one measure run. Output to a pickle and xlsx file.

    :param run_dir: Directory with log file and directories for scope and spectra data.
    :param reactor: The used reactor
    :param spect_dir: subdirectory with the spectra data.
    :param scope_dir: subdirectory with the scope data
    :param log_file: name of the logfile with input parameters [voltage, freq, v18,9ohm input, spectfile, Temp, airflow]
    :param scope_file_name_index: column of log_file that is used for the scope filenames, 0=volt, 1=freq, 2=pulsew
    :param meas: lenght of the measure cell
    :param current_scaling: Scale of the current. 20 for red sensor, 100 for green sensor
    :param delay: Delay for the first scope line, to align lines.
    :param voltage_offset: Zero point for the voltage in the easyscope files. Used if scope line is not at div zero.
    :param current_offset: Zero point for the current in the easyscope files. Used if scope line is not at div zero.
    :param waveform_loose_stability: Set to True if the waveforms are very bad, this sets the stability checking less strict. Only used with scope_multiple=true
    :return: none
    """
    print("Calc run for "+run_dir)
    if log_file not in os.listdir(run_dir):
        return
    assert spect_dir in os.listdir(run_dir)
    assert scope_dir in os.listdir(run_dir)
    # Get ozone concentrations:
    all_ozone = ozone_concentration(path_name=run_dir + '/' + spect_dir, opt_path=meas)
    all_ppm = ozone_ppm(all_ozone)

    # assuming format: voltage, freq, voltage on 18,9ohm input, meting spect, Temp, airflow
    log_file = run_dir + '/' + log_file
    resistor_val = 18.9  # resistance of measure resistor for input current

    sheet = load_workbook(filename=log_file, read_only=True).active
    data = []
    for row in sheet.iter_rows(min_row=2):
        data_row = []
        for cell in row:
            # read values from workbook
            data_row.append(cell.value)

        if all([not data for data in data_row[0:3]]):  # if line is all empty
            print('finished reading')
            break
        if None in data_row[0:7]:
            print("Error! Some empty cells found in: " + str(data_row))
        # 0: voltage        [V]
        # 1: freq           [Hz]
        # 2: duration       [us]
        # 3: v18, 9ohm input[V]
        # 4: spectfile      [filename]
        # 5: Temp           [deg. Cels.]
        # 6: airflow        [ls/min]
        # 7: inductance     [uH]
        input_i = data_row[3] / resistor_val  # generator input current [A]
        input_p = input_i * data_row[0]  # generator input power [W]
        input_p_k = input_p / 3.6e6  # input power in kWh/s
        input_e_pulse = input_p / data_row[1]  # input energy per pulse [J]
        co3 = all_ozone[data_row[4]]  # mol/m3,
        co3g = co3 * 48  # gram/m3 o3
        lss = data_row[6] / 60  # liter air per second
        m3s = lss / 1000  # ls/min/1000/60=m3/s
        o3f = co3g * m3s  # (gram/m3)*(m3/s)=gram/second o3

        if len(data_row) >= 8:
            if reactor == REACTOR_GLASS_LONG:
                if data_row[7]:
                    assert data_row[7] in INDUCTANCE_LONG_REACTOR  # 26 uH coil with long glass reactor
                else:
                    data_row[7] = 26  # uH
            else:
                assert data_row[7] in INDUCTANCE_SHORT_REACTOR  # valid values for inductance
        elif len(data_row) == 7:  # inductance not supplied
            data_row.append(0)
            if reactor == REACTOR_GLASS_LONG:
                data_row[7] = 26  # 26 uH coil with long glass reactor
        else:
            raise Exception
        assert np.isfinite(data_row[7])
        energy_loose_stability = (data_row[7] != 0)  # if using coil, energy calculation is less stable.
        # get output waveforms
        dic = {}
        try:
            if scope_multiple:
                print('input csv: ' + str(data_row[scope_file_name_index]))
                lines = get_vol_cur_multiple(run_dir + '/' + scope_dir + '/' + str(data_row[scope_file_name_index]),
                                             current_scaling=current_scaling,
                                             delay=delay, voltage_offset=voltage_offset, current_offset=current_offset,
                                             )
                assert any(lines)
                output_calc = calc_output_avg(lines, gen_res_high=225, gen_res_low=50, loose_stability=waveform_loose_stability, energy_loose_stability=energy_loose_stability)
            else:
                line = get_vol_cur_single(run_dir + '/' + scope_dir + '/' + str(data_row[scope_file_name_index]),
                                          current_scaling=current_scaling,
                                          delay=delay, voltage_offset=voltage_offset, current_offset=current_offset)
                assert line
                # calculate output parameters from d_calc.py and append them to the dict with prepend '_output'
                output_calc = calc_output(line, gen_res_high=225, gen_res_low=50)

            # append calculated output values to the dict, this is what ends up in the pickle and excel files.
            for key, val in output_calc.items():
                dic['output_' + key] = val
            # output power on plasma [Watt]
            output_p_plasma = (dic['output_e_plasma'] * data_row[1])
            output_p_plasma_single = np.array(dic['output_e_plasma_single']) * data_row[1]
        except IOError:
            output_p_plasma = output_p_plasma_single = 0

        dic = {**dic,
               # manually noted values (inputs)
               'input_v': data_row[0],  # input voltage in V
               'input_v_output': data_row[0] * 15,  # excpected output voltage of generator
               'input_f': data_row[1],  # pulse frequency in Hz
               'input_l': data_row[2],  # pulse length in us.
               'temperature': data_row[5] or 22,  # temperature in deg Celsius, default to 22
               'airflow_lm': data_row[6],  # airflow in ls/min
               'airflow_ls': lss,  # airflow in ls/s
               'airflow_m3s': m3s,  # airflow in m3/s
               'inductance': data_row[7],
               'reactor': reactor,

               # calculated values from noted values. (This is not compensated for resistive losses in generator!!)
               'input_c': input_i,  # input current to generator
               'input_p': input_p,  # input power to generator
               'input_e_pulse': input_e_pulse,  # input energy per pulse in Joule
               'input_energy_dens': input_p / lss,  # input energy to generator in Joule/ Liter used air

               # values from ozone spectrum, as obtained from spectrometer.
               'o3_gramm3': co3g,
               'o3_molm3': co3,  # o3 concentration in Mol/m3
               'o3_ppm': all_ppm[data_row[4]],  # o3 concentration in ppm
               'o3_gramsec': o3f,  # o3 production in gram/second

               # o3 generation efficiency
               'input_yield_gj': o3f / input_p if input_p else 0,  # efficiency in gram/Joule
               'input_yield_gkwh': o3f / input_p_k if input_p_k else 0,  # efficiency in gram/kWh
               'output_p_avg': output_p_plasma,
               'output_energy_dens': output_p_plasma / lss,
               'output_yield_gj': o3f / output_p_plasma if output_p_plasma else 0,
               'output_yield_gkwh': o3f / (output_p_plasma / 3.6e6) if output_p_plasma else 0,
               'output_energy_dens_single': output_p_plasma_single / lss,
               'output_yield_gj_single': o3f / output_p_plasma_single if any(output_p_plasma_single) else 0,
               'output_yield_gkwh_single': o3f / (output_p_plasma_single / 3.6e6) if any(output_p_plasma_single) else 0,
               'e_eff': output_p_plasma / input_p if output_p_plasma else 0,
               }

        # add this measurement to the total list.
        data.append(dic)

    # sort data
    if scope_file_name_index == 0:
        data = sort_data(data, 'input_v')
    elif scope_file_name_index == 1:
        data = sort_data(data, 'input_f')
    elif scope_file_name_index == 2:
        data = sort_data(data, 'input_l')
    elif scope_file_name_index in [4,8]:
        pass
        # data = sort_data(data, 'input_l')
    else:
        raise Exception('Invalid scope file name index')

    # save pickle
    with open(run_dir + '/data.pkl', 'wb') as f:
        pickle.dump(data, f, )

    # save xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    # make header row.
    keys = []
    try:  # row 5 probably contains all headers, since the first few will miss some as they are reference lines.
        ref_line = data[4]
    except:  # if row 5 does not exists, for very short runs.
        ref_line = data[0]
    for key, value in sorted(ref_line.items()):
        # don't save arrays of values, it is too much information
        if type(value) in [float, int] or type(value).__module__ == 'numpy': # all numeric values.
            try:
                value = float(value)
                if not np.isfinite(value):
                    continue
            except:
                # skip columns with arrays as value.
                continue
        elif type(value) is not str:  # all except strings, like lists and arrays, are discarded for excel data.
            continue
        keys.append(key)
    ws.append(keys)

    # fill data for all keys
    for meas in data:
        row = []
        for key in keys:
            value = meas[key]
            if type(value) is not str:
                try:
                    if not np.isfinite(value):  # replace inf and -inf with 0
                        value = 0
                except:
                    value = 0
            row.append(value)
        ws.append(row)
    # make unique filename because excel cannot open multiple workbooks with the same name
    file_name = '-'.join(run_dir.split('/')[-2:]).strip().strip('-')
    wb.save(filename=run_dir + '/' + file_name + '.xlsx')

    print("finished writing")
    return 1


if __name__ == '__main__':
    # path = "G:/Prive/MIJN-Documenten/TU/62-Stage/20180104-500hz/" # directory with subdirectories with measurements
    # path = "G:/Prive/MIJN-Documenten/TU/62-Stage/20180104-500hz/"  # directory with subdirectories with measurements
    # path = "G:/Prive/MIJN-Documenten/TU/62-Stage/20180111/"  # directory with subdirectories with measurements
    # path = "G:/Prive/MIJN-Documenten/TU/62-Stage/20180109/"  # directory with subdirectories with measurementspath = "G:/Prive/MIJN-Documenten/TU/62-Stage/20180103-1000Hz/"  # directory with subdirectories with measurements
    ### to run dir with subdirs:

    dirs = glob.glob(path+'/run*')
    ### to run one dir
    # dirs = ['120171229']
    # dirs = ['run2-1us-q']
    # length of used measure cell
    meas_len = SHORT_MEAS_LEN
    # capacitance of used reactor
    react_cap = REACTOR_GLASS_SHORT_QUAD
    # which column of log.xlsx contains the filename for scope. 0=volt, 1=freq, 2=pulsew
    scope_file_name_index = 0
    # whether multiple scope spectra are stored for each measurement. If true, save as xxx_y.csv with y as index number
    scope_multiple = True
    # scaling for current sensor is not done in scope, do it manually
    current_scaling = 0.5  # 0.5 for red current probe in v-range, -0.1 for pearson (inverted) in v-range, -100 for mv range.
    # compensate for delay in line, in array index (=usually 1ns)
    delay = 0
    for dir in dirs:
        run_dir = dir + '/'
        run_dir = run_dir.replace('\\', '/').replace('//', '/')
        if os.path.isdir(run_dir):
            print(run_dir)
            calc_run(run_dir,
                     meas=meas_len,
                     scope_file_name_index=scope_file_name_index,
                     scope_multiple=scope_multiple,
                     reactor=react_cap,
                     current_scaling=current_scaling,
                     delay=delay)
        else:
            print('Path ' + str(run_dir) + ' does not exist.')
