"""
Parse log excel file
Get ozone concentration
Get scope plots

Make list of dicts. Each dict is one measurement and contains all relevant input and output values.
Jeroen van Oorschot 2017-2018, Eindhoven University of Technology
jeroen@jjvanoorschot.nl
"""
import numpy as np
import pickle
from openpyxl.reader.excel import load_workbook
from scipy import integrate

from scope_parse.c_get_lines import get_vol_cur_single
from spectrum_parse.c_concentration import ozone_concentration, ozone_ppm

# Folder with measuremtn. Folder has to contain:
# * log.txt
# * spect/m00000.txt  with spectral data
# * scope/000.csv with 000=used input voltage
run_dir = "G:/Prive/MIJN-Documenten/TU/62-Stage/20171229"
spect_dir = '/spect'
scope_dir = '/scope'
log_file = 'log.xlsx'  # must be xlsx with at least [voltage, freq, v18,9ohm input, spectfile, Temp, airflow]

######################
# script starts here #
######################

# Get ozone concentrations:
all_ozone = ozone_concentration(path_name=run_dir+spect_dir)
all_ppm = ozone_ppm(all_ozone)

# assuming format: voltage, freq, voltage on 18,9ohm input, meting spect, Temp, airflow
log_file = run_dir + '/' + log_file
resistor_val = 18.9 # resistance of measure resistor for input current

sheet = load_workbook(filename=log_file, read_only=True).active
data = []
for row in sheet.iter_rows(min_row=2):
    data_row = []
    for cell in row[0:7]:
        # read values from workbook
        data_row.append(cell.value)
    if all([not data for data in data_row]):  # if line is all empty
        print('finished reading')
        break
    if None in data_row:
        print("Error! Some empty cells found in: "+str(data_row))
    I = data_row[3]/resistor_val  # generator input current [A]
    P = I*data_row[0]  # generator input power [W]
    Pk = P /1000 * 3600  # input power in kWh/s
    iEp = P/data_row[1]  # input energy per pulse [J]
    #iEpk = iEp / 1000 * 3600 # input energy per pulse {kWh]
    co3 = all_ozone[data_row[4]]  # mol/m3
    co3g = co3 * 48 # gram/m3 o3
    lss = data_row[6]/60 # liter air per second
    m3s = lss/1000 # ls/min/1000/60=m3/s
    o3f = co3g/m3s  # (gram/m3)*(m3/s)=gram/second o3

    # get output waveforms
    try:
        output_time, output_v, output_i = get_vol_cur_single(run_dir+scope_dir+'/'+str(data_row[0])+'.csv')
        output_p = output_i * output_v
        integrate.cumtrapz(output_p, output_time, initial=0)
    except IOError:
        output_time = output_v = output_i = output_p = output_e = 0

    dic = {
        # manually noted values (inputs)
        'input_voltage': data_row[0],   # input voltage in V
        'pulse_frequency': data_row[1], # pulse frequency in Hz
        'pulse_duration': data_row[2],  # pulse duration in us.
        'temperature': data_row[5],     # temperature in deg Celsius
        'airflow': data_row[6],         # airflow in ls/min

        # calculated values from noted values. (This is not compensated for resistive losses in generator!!)
        'input_current': I,             # input current to generator
        'input_power': P,               # input power to generator
        'input_pulse_energy': iEp,      # input energy per pulse in Joule
        'input_energy_dens': P / lss,   # input energy to generator in Joule/ Liter used air

        # values from ozone spectrum, as obtained from spectrometer.
        'o3_concentration': co3,        # o3 concentration in Mol/m3
        'o3_ppm': all_ppm[data_row[4]], # o3 concentration in ppm
        'o3_gramsec': o3f,              # o3 production in gram/second

        # scope spectra
        'output_time': output_time,
        'output_voltage': output_v,
        'output_current': output_i,
        'output_power': output_p,
        'output_energy': output_e,

        # o3 generation efficiency
        'input_eff_gj': o3f/P,          # efficiency in gram/Joule
        'input_eff_gkwh': o3f/Pk,       # efficiency in gram/kWh

    }
    data.append(dic)

with open(run_dir+'/data.pkl', 'wb') as f:
    pickle.dump(data, f, )